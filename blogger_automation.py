from tavily import TavilyClient
import google.generativeai as genai
from config.settings import GEMINI_API_KEY,TAVILY_API_KEY #Make settings.py file in config folder and save the api keys with these variable name
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
import time
import os
import re

# Initialize Clients
client_tavily = TavilyClient(TAVILY_API_KEY)

genai.configure(api_key=GEMINI_API_KEY)



def scrape_urls(url_list):
# To install: pip install tavily-python
   from tavily import TavilyClient
   print(url_list)
   client = TavilyClient("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")
   response = client.extract(
   urls=url_list,
   extract_depth="advanced")
   return response


def generate_gemini_response(prompt, context,images):
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(f'''You are a finance and credit card blog writer. Generate a detailed, SEO-optimized HTML blog on "{prompt}" with:

Complete details (features, benefits, drawbacks, only from context).
Comparisons & Tables (fees, rewards, spending breakdown if applicable).
Use Cases (real-world examples, cashback/points breakdown in a table).
SEO-friendly Title (clear, engaging, relevant).
Affiliate Links: Hyperlink product names & add "Apply Now" buttons.
Semantic, Responsive HTML (headings, paragraphs, tables, images, figcaptions).
Tables:
Responsive with horizontal scroll.
White headings, black background.
Prevent word breaks in cells.
Images: Include at least 2 images with credit.
Category Label: Add <label> at the end with only the product category (e.g., "Credit Card", "Comparison").
Readable & Engaging: Short paragraphs, bullet points, min. 800 words.
Strict Data Usage: Do not add any extra information beyond the provided context.
🔥 Exclusive Deals:

Card Wallet Offer [https://amzn.to/4iI9FqV]

Flipkart Axis Bank Credit Card Pre-Approval [https://dl.flipkart.com/s/6auzIbNNNN]
Structure - 
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Blog Post</title>
    <style>
        body ( font-family: Arial, sans-serif; margin: 20px; background-color: #121212; color: #fff; )
        h1, h2, h3 ( color: white; )
        .content ( max-width: 800px; margin: auto; )
        .image-container ( text-align: center; margin: 20px 0; )
        .image-container img ( max-width: 100%; height: auto; )
        .image-credit ( font-size: 12px; color: #bbb; )
        .table-container ( overflow-x: auto; margin-top: 20px; )
        table ( width: 100%; border-collapse: collapse; white-space: nowrap; )
        th, td ( padding: 10px; border: 1px solid #fff; text-align: left; )
        th ( background-color: black; color: white; )
        .buy-now ( display: block; text-align: center; margin: 20px 0; padding: 10px; background: red; color: #fff; text-decoration: none; )
        .category-label ( display: block; margin-top: 20px; font-size: 16px; font-weight: bold; )
    </style>
</head>
<body>
    <div class="content">
        <h1>Main Heading</h1>
        <p>Introduction paragraph.</p>
        
        <h2>Subheading</h2>
        <p>Descriptive text.</p>
        
        <div class="image-container">
            <img src="image1.jpg" alt="Image 1">
            <p class="image-credit">Image Credit: Source 1</p>
        </div>

        <h2>Data Table</h2>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>Feature</th>
                        <th>Details</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>Specification 1</td>
                        <td>Value 1</td>
                    </tr>
                    <tr>
                        <td>Specification 2</td>
                        <td>Value 2</td>
                    </tr>
                </tbody>
            </table>
        </div>

    </div>
                                          
    <p>🔥 <strong>Exclusive Deal Alert!</strong> 🔥</p>
    <p>Upgrade your style with this sleek and secure <strong>Card Wallet</strong>! 💳✨</p>
    <p>👉 <a href="https://amzn.to/4iI9FqV" class="buy-now">Grab Yours Now! 💥</a></p>

    <p>🔥 <strong>Exclusive Deal Alert!</strong> 🔥</p>
    <p>Get the <strong>Flipkart Axis Bank Credit Card</strong> and earn rewards! 🚀💳</p>
    <p>👉 <a href="https://dl.flipkart.com/s/6auzIbNNNN" style="color: #ffffff; background-color: #ff6600; padding: 10px 15px; border-radius: 5px; text-decoration: none; font-weight: bold; display: inline-block;">Claim Your Card Now! 💥</a></p>

    <label class="category-label">Product Category</label>


</body>
</html>
 
**Additional Info:**  
Context: {context}  
Image URLs: {images}  
''')
        return response.text
    except Exception as e:
        print("Gemini failed, falling back to Groq:", str(e))
        return "Gemini failed, falling back to Groq:"



def search_web(query):
    response = client_tavily.search(
        query=query,
        max_results=4,
        include_images=True,
        exclude_domains=["youtube.com"],
        include_raw_content=True,
        include_image_descriptions=False
    )
    return response

def generate_blog(query):
    # Fetch web data
    search_results = search_web(query)
    image_urls = search_results["images"]
    result_dict = {i: result for i, result in enumerate(search_results["results"], start=1)}
    urls = [item["url"] for item in result_dict.values()]
    data=scrape_urls(urls)
    # Generate blog content using LLMs (Gemini primary, Groq fallback)
    blog_html = generate_gemini_response(query, data, image_urls)
    
    return blog_html


def automate_blogger_post(title, blog_html,label):
    # Set Edge options to use the default profile
    options = Options()
    options.add_argument("user-data-dir=C:\\Users\\vinay\\AppData\\Local\\Microsoft\\Edge\\User Data")  # add yours path this is example
    options.add_argument("profile-directory=Default")  # Use your main profile
    options.add_argument("--start-maximized")

    # Initialize WebDriver
    driver = webdriver.Edge(options=options)

    try:
        # Open Blogger posts page
        driver.get("https://www.blogger.com/blog/posts/858567687044551803?hl=en")
        time.sleep(5)  # Wait for the page to load

        # Click on the "Posts" link
        posts_link = driver.find_element(By.XPATH, "//*[@id='yDmH0d']/c-wiz/div[1]/gm-raised-drawer/div/div[2]/div/c-wiz/div[3]/div/div")
        posts_link.click()
        time.sleep(5)

        # Enter "title" in the input field
        input_field = driver.find_element(By.XPATH, "//*[@id='yDmH0d']/c-wiz[2]/div/c-wiz/div/div[1]/div[1]/div[1]/div/div[1]/input")
        input_field.clear()  # Clear existing text if needed
        input_field.send_keys(title)
        time.sleep(1)

        inputlabel=driver.find_element(By.XPATH, '/html/body/div[8]/c-wiz[2]/div/c-wiz/div/div[2]/div/div/div[4]/span/c-wiz/div/div[2]/div[1]/span/div/div[1]/div[1]/div[2]/textarea')
        inputlabel.clear()
        inputlabel.send_keys(label)
        time.sleep(1)

        # Locate CodeMirror editor and enter text
        editor = driver.find_element(By.CLASS_NAME, "CodeMirror-scroll")
        driver.execute_script("arguments[0].click();", editor)
        time.sleep(1)

        # Enter text using JavaScript (since send_keys might not work)
        script = f"""
        var editor = document.querySelector('.CodeMirror').CodeMirror;
        editor.setValue({repr(blog_html)});
        """
        driver.execute_script(script)
        time.sleep(5)

        # Click Preview button
        preview_button = driver.find_element(By.XPATH, "//div[@role='button' and @aria-label='Preview post']")
        driver.execute_script("arguments[0].click();", preview_button)  # Ensures click works
        time.sleep(12)

        # Switch to new tab and close it after 2 seconds
        driver.switch_to.window(driver.window_handles[1]) 
        time.sleep(2)
        driver.execute_script("window.close();")

        # Switch back to the main tab
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(2)

        # Click Publish button
        publish_button = driver.find_element(By.XPATH, "//*[@id='yDmH0d']/c-wiz[2]/div/c-wiz/div/div[1]/div[2]/div[4]")
        publish_button.click()
        time.sleep(2)

        # Confirm publish
        confirm = driver.find_element(By.XPATH, "//*[@id='yDmH0d']/div[4]/div/div[2]/div[3]/div[2]")
        confirm.click()
        time.sleep(5)

    finally:
        driver.quit()
import openpyxl

def get_next_topic(file_name="topics.xlsx"):

    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active  # Use the first sheet

        # Find the "RateMyCard" column
        rate_my_card_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "RateMyCard":
                rate_my_card_col = col
                break
        
        if rate_my_card_col is None:
            raise ValueError("Column 'RateMyCard' not found in the Excel file.")

        # Find or create the "Used" column
        used_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "Used":
                used_col = col
                break

        if used_col is None:
            used_col = sheet.max_column + 1
            sheet.cell(row=1, column=used_col, value="Used")  # Add header

        # Find the next unused topic
        next_topic = None
        for row in range(2, sheet.max_row + 1):  # Start from second row (excluding headers)
            used_cell = sheet.cell(row=row, column=used_col)
            if used_cell.value is None or used_cell.value is False:
                next_topic = sheet.cell(row=row, column=rate_my_card_col).value
                used_cell.value = True  # Mark as used
                break

        # If all topics are used, reset and start over
        if next_topic is None:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=used_col, value=False)  # Reset all to unused
            next_topic = sheet.cell(row=2, column=rate_my_card_col).value
            sheet.cell(row=2, column=used_col, value=True)  # Mark first topic as used

        # Save the updated file
        workbook.save(file_name)
        workbook.close()  # Close workbook

        return next_topic

    except PermissionError:
        print("Error: Unable to save the file. Please close it if it's open in another program.")
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None

# Example Usage
next_topic = get_next_topic()
if next_topic is None:
    print("Script stopped due to an error.")
    exit(1)  # Stop execution with a non-zero exit code (indicates error)
query = next_topic
blog_html = generate_blog(query)
soup = BeautifulSoup(blog_html, "html.parser")
label=soup.label.string
blog_html = re.sub(r"<label[^>]*>.*?</label>", "", blog_html, flags=re.DOTALL)
blog_html = re.sub(r"<h1>.*?</h1>", "", blog_html, flags=re.DOTALL)
title = soup.title.string
blog_html = blog_html.replace("```html", "").replace("```", "").strip()
os.system("taskkill /F /IM msedge.exe")
automate_blogger_post(title, blog_html,label)
